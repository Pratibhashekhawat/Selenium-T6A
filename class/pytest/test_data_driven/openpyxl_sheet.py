import openpyxl
wb=openpyxl.Workbook()
sheetName="Sheet1"
if sheetName in wb.sheetnames:
    ws=wb[sheetName]
else:
    ws=wb.create_sheet(sheetName)
ws['A1']="user"
ws['B1']="password"

ws.append(['standard_user','secret_sauce'])
ws.append(['problem_user','secret_sauce'])
# ws.append(['user3','789'])
wb.save('sample.xlsx')

for row in ws.iter_rows(values_only=True):
    print (row)



